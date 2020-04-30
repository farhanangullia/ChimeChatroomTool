from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from time import sleep
import configparser
from tqdm import tqdm

config = configparser.ConfigParser()
config.read('config/parameters.ini')
CHATROOM_URL = config['CHIME']['CHATROOM_URL']
EMAIL_LIST_FILE = config['PATH']['EMAIL_LIST_FILE']
SHEET_NAME = config['PATH']['SHEET_NAME']
CHROME_DRIVER_PATH = config['PATH']['CHROME_DRIVER_PATH']
options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')
driver = webdriver.Chrome(options=options, executable_path=CHROME_DRIVER_PATH) 
email_list_wb = openpyxl.load_workbook(EMAIL_LIST_FILE)

def start_chime_app():
    driver.get(CHATROOM_URL)
    input("Login to your Chime account on the browser then press Enter here to start...")

def add_members():
    failed_members = []
    added_members = []
    ws = email_list_wb[SHEET_NAME]
    try:
        print("Adding members to chatroom from excel sheet...")
        for cell in tqdm(ws['A']):
            email = cell.value.strip()
            add_member_url = "{}/members/add?email={}".format(CHATROOM_URL, email)
            driver.get(add_member_url)
            sleep(3)
            try:
                element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//*[@class='ContactListItem__button']"))) # contact button in list
                memberName = element.get_attribute("title")
                element.click()
                added_members.append(str(memberName))
                sleep(1)
                element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//*[@class='Button Button__primary' and @type='submit']"))) # add button
                element.click()
            except Exception as e:
                print("Unable to find contact: " + email + " Exception: " + str(e))
                failed_members.append(email)
    except Exception as e:
        print("Error adding member: " + str(e))
    update_excel(failed_members, added_members)

def update_excel(failed_members, added_members):
    try:
        print("Updating excel file...")
        failed_members_sheet = email_list_wb.create_sheet("Failed")
        added_members_sheet = email_list_wb.create_sheet("Added")
        failed_members_sheet.append(failed_members)
        added_members_sheet.append(added_members)
        email_list_wb.save(filename = EMAIL_LIST_FILE)
        print("File saved...")
    except Exception as e:
        print("Error updating excel file: " + str(e))

if __name__ == "__main__":
    start_chime_app()
    add_members()